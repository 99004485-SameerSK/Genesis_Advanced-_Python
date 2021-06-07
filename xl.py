"""openpyxl for handling xl files"""
from openpyxl import load_workbook
import openpyxl

class Details:
    """For having details of each PSNum"""
    def __init__(self, name):
        """Constructor"""
        self.name = name
    def fun(self):
        """Takes values from Excel file"""
        for listt in range(0,len(self.name)):
            q.append(self.name[listt])

    @classmethod
    def write1(self):
        """Writes to a new Excel file"""
        q.insert(0, psnum)
        print(q)
        try:
            if len(q)==len(p):
                index=1
            assert index==1
        except IndexError:
            print("no match in details number")
        else:
            workbok = openpyxl.Workbook()
            sheet = workbok.active
            for items in range(0,len(p)):
                fun_xl = sheet.cell(row=1, column=items+1)
                fun_xl.value = p[items]
            workbok.save("demo.xlsx")
            for detail in range(0,len(q)):
                fun_xl = sheet.cell(row=2, column=detail+1)
                fun_xl.value = q[detail]
            workbok.save("demo.xlsx")
            print("successfully entered")

p=["PSNUM","I","II","III","IV","V","VI","VII","V","CITY","LANGUAGE","AREA","YEAR","HOBBY"]
print("Enter PS Number for details")
q=[]
try:
    psnum = int(input())
    Workbook = load_workbook(filename="Book1.xlsx")
    a = Workbook.sheetnames
    for i,a in enumerate(a):
        Datasheet = Workbook[a[i]]
        for j in range(1, 17):
            aa = Datasheet.cell(row=j, column=1).value
            if aa == psnum:
                x = j
        for value in Datasheet.iter_rows(min_row=x, max_row=x, min_col=2, values_only=True):
            y = Details(value)
            y.fun()
    y.write1()
except NameError:
    print("Error no match in ps number")
